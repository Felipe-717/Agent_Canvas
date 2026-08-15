"""El chat: conversar con normalidad, adjuntar archivos, iterar sobre graficos.

Lo que se comprueba no es que el modelo acierte, sino que la conversacion
recuerde, que los archivos sean adjuntos y no un paso previo, y que de un
grafico se guarde su especificacion y nunca sus numeros.
"""

from __future__ import annotations

from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from agentcanvas.application.ports.llm import Role
from agentcanvas.bootstrap.container import Container
from agentcanvas.domain.chat.entities import MessageRole
from agentcanvas.infrastructure.persistence.conversations import (
    SqlAlchemyConversationRepository,
)
from agentcanvas.infrastructure.web.app import create_app
from tests.fakes import FakeLLM, text_response, tool_response

VENTAS_POR_REGION: dict[str, Any] = {
    "type": "bar",
    "title": "Ventas por región",
    "x": {"field": "region"},
    "y": [{"field": "valor", "aggregation": "sum"}],
}
DE_BARRAS_A_TARTA: dict[str, Any] = {**VENTAS_POR_REGION, "type": "pie"}


@pytest.fixture
async def client(container: Container) -> AsyncIterator[AsyncClient]:
    app = create_app(container)
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


@pytest.fixture
def excel(tmp_path: Path) -> bytes:
    """Un archivo con la cabecera lejos de la primera fila."""
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Ventas"
    sheet["A1"] = "Informe confidencial"
    for column, name in enumerate(["region", "valor"], start=1):
        sheet.cell(row=3, column=column, value=name)
    for offset, (region, valor) in enumerate([("Norte", 100), ("Sur", 150), ("Norte", 20)]):
        sheet.cell(row=4 + offset, column=1, value=region)
        sheet.cell(row=4 + offset, column=2, value=valor)
    path = tmp_path / "ventas.xlsx"
    book.save(path)
    return path.read_bytes()


async def _open(client: AsyncClient) -> str:
    response = await client.post("/api/conversations")
    assert response.status_code == 201, response.text
    conversation_id: str = response.json()["id"]
    return conversation_id


async def _send(
    client: AsyncClient, conversation: str, text: str, upload: bytes | None = None
) -> dict[str, Any]:
    files = {"file": ("ventas.xlsx", upload, "application/vnd.ms-excel")} if upload else None
    response = await client.post(
        f"/api/conversations/{conversation}/messages", data={"text": text}, files=files
    )
    assert response.status_code == 200, response.text
    payload: dict[str, Any] = response.json()
    return payload


# ---------------------------------------------------------- conversar sin datos


async def test_a_plain_question_gets_a_plain_answer(
    client: AsyncClient, llm: FakeLLM
) -> None:
    # Sin archivos de por medio: es un chat, no un formulario.
    conversation = await _open(client)
    llm.queue(text_response("Puedo ayudarte a analizar hojas de cálculo."))

    turn = await _send(client, conversation, "hola, ¿qué sabes hacer?")

    assert turn["assistant_message"]["text"].startswith("Puedo ayudarte")
    assert turn["assistant_message"]["artifacts"] == []


async def test_the_conversation_takes_its_title_from_the_first_message(
    client: AsyncClient, llm: FakeLLM
) -> None:
    conversation = await _open(client)
    llm.queue(text_response("Claro."))

    await _send(client, conversation, "analiza las ventas del trimestre")

    listed = (await client.get("/api/conversations")).json()
    # Pedirle un titulo al modelo costaria una llamada por algo ya escrito.
    assert listed[0]["title"] == "analiza las ventas del trimestre"


async def test_the_model_receives_the_whole_history(
    client: AsyncClient, llm: FakeLLM
) -> None:
    conversation = await _open(client)
    llm.queue(text_response("Hola."), text_response("Sí, lo recuerdo."))

    await _send(client, conversation, "me llamo Felipe")
    await _send(client, conversation, "¿cómo me llamo?")

    # El segundo turno tiene que llevar el primero: sin esto no hay memoria.
    contents = [message.content for message in llm.requests[1].messages]
    assert any("Felipe" in content for content in contents)


# ------------------------------------------------------------ adjuntar archivos


async def test_a_file_travels_with_the_message(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation = await _open(client)
    llm.queue(text_response("He recibido el archivo."))

    turn = await _send(client, conversation, "mira esto", upload=excel)

    attachments = turn["user_message"]["attachments"]
    assert [a["filename"] for a in attachments] == ["ventas.xlsx"]
    # Y el modelo se entera de que hay un archivo y con que id.
    assert "ventas.xlsx" in llm.requests[0].messages[-1].content


async def test_preparing_a_table_produces_a_dataset_artifact(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation = await _open(client)
    llm.queue(text_response("Voy a mirarlo."))
    turn = await _send(client, conversation, "prepáralo", upload=excel)
    file_id = turn["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response(
            "preparar_datos",
            {"archivo": file_id, "hoja": "Ventas", "fila_cabecera": 3, "nombre": "ventas"},
        ),
        text_response("Listo, he preparado la tabla."),
    )
    turn = await _send(client, conversation, "usa la hoja Ventas, cabecera en la fila 3")

    artifacts = turn["assistant_message"]["artifacts"]
    assert artifacts[0]["kind"] == "dataset"
    assert artifacts[0]["row_count"] == 3
    assert artifacts[0]["columns"] == ["region", "valor"]
    # La procedencia se guarda en palabras, para poder ensenarsela al usuario.
    assert "fila 3" in artifacts[0]["origin"]


# ------------------------------------------------------------ graficos en el chat


async def _with_dataset(client: AsyncClient, llm: FakeLLM, excel: bytes) -> tuple[str, str]:
    conversation = await _open(client)
    llm.queue(text_response("Recibido."))
    turn = await _send(client, conversation, "toma", upload=excel)
    file_id = turn["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response(
            "preparar_datos", {"archivo": file_id, "hoja": "Ventas", "fila_cabecera": 3}
        ),
        text_response("Preparado."),
    )
    turn = await _send(client, conversation, "prepáralo")
    return conversation, turn["assistant_message"]["artifacts"][0]["dataset_id"]


async def test_a_chart_appears_inside_the_conversation(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, dataset_id = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": dataset_id, "especificacion": VENTAS_POR_REGION}
        ),
        text_response("Aquí tienes las ventas por región."),
    )

    turn = await _send(client, conversation, "dibuja las ventas por región")

    artifact = turn["assistant_message"]["artifacts"][0]
    assert artifact["kind"] == "visual"
    assert artifact["spec"]["type"] == "bar"
    assert artifact["data"]["rows"] == [
        {"region": "Norte", "sum_valor": 120.0},
        {"region": "Sur", "sum_valor": 150.0},
    ]


async def test_the_chart_data_is_recomputed_not_stored(
    client: AsyncClient, llm: FakeLLM, excel: bytes, container: Container
) -> None:
    conversation, dataset_id = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": dataset_id, "especificacion": VENTAS_POR_REGION}
        ),
        text_response("Ahí va."),
    )
    await _send(client, conversation, "dibuja")

    session = container.session_factory()
    try:
        messages = await SqlAlchemyConversationRepository(session).list_messages(conversation)
    finally:
        await session.close()

    visual = messages[-1].visuals[0]
    # Lo guardado es la especificacion. Los numeros del endpoint se acaban de
    # calcular: por eso una conversacion de hace un mes muestra datos de hoy.
    assert visual.spec.title == "Ventas por región"
    assert not hasattr(visual, "data")

    history = (await client.get(f"/api/conversations/{conversation}/messages")).json()
    assert history[-1]["artifacts"][0]["data"]["rows"]


async def test_an_impossible_chart_is_explained_to_the_model(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, dataset_id = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "crear_visual",
            {
                "dataset_id": dataset_id,
                "especificacion": {**VENTAS_POR_REGION, "x": {"field": "departamento"}},
            },
        ),
        text_response("Perdona, esa columna no existe."),
    )

    await _send(client, conversation, "dibuja por departamento")

    feedback = next(
        message
        for message in llm.requests[-1].messages
        if message.content.startswith("No se puede dibujar")
    )
    assert "departamento" in feedback.content
    assert "region" in feedback.content


async def test_asking_for_a_chart_without_data_points_to_preparing_it(
    client: AsyncClient, llm: FakeLLM
) -> None:
    conversation = await _open(client)
    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": "ds_inexistente", "especificacion": VENTAS_POR_REGION}
        ),
        text_response("Necesito que me pases un archivo primero."),
    )

    await _send(client, conversation, "dibuja las ventas")

    feedback = next(
        message for message in llm.requests[-1].messages if "ds_inexistente" in message.content
    )
    assert "preparar_datos" in feedback.content


# ------------------------------------------------------------------- memoria


async def test_iterating_on_a_chart_works_because_the_history_travels(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, dataset_id = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": dataset_id, "especificacion": VENTAS_POR_REGION}
        ),
        text_response("Aquí tienes."),
    )
    await _send(client, conversation, "ventas por región")

    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": dataset_id, "especificacion": DE_BARRAS_A_TARTA}
        ),
        text_response("Cambiada a tarta."),
    )
    turn = await _send(client, conversation, "ahora hazla de tarta")

    # "la" es resoluble porque el turno anterior viaja en el historial.
    previous = [m.content for m in llm.requests[-2].messages]
    assert any("Ventas por región" in content for content in previous)
    assert turn["assistant_message"]["artifacts"][0]["spec"]["type"] == "pie"


async def test_the_history_summarises_artifacts_instead_of_resending_data(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, dataset_id = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "crear_visual", {"dataset_id": dataset_id, "especificacion": VENTAS_POR_REGION}
        ),
        text_response("Listo."),
    )
    await _send(client, conversation, "dibuja")

    llm.queue(text_response("Sí."))
    await _send(client, conversation, "¿sigue ahí?")

    # Sin el prompt del sistema, que menciona claves de ejemplo.
    history = "\n".join(
        m.content for m in llm.requests[-1].messages if m.role is not Role.SYSTEM
    )
    # Reenviar los numeros de cada grafico multiplicaria el coste sin anadir
    # informacion: el modelo ya sabe lo que dibujo.
    assert "dibujado el gráfico" in history or "dibujado el grafico" in history
    assert "sum_valor" not in history


# ------------------------------------------------------------------ lienzos


async def test_a_canvas_can_mix_several_sources(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, first = await _with_dataset(client, llm, excel)
    llm.queue(
        tool_response(
            "preparar_datos",
            {"archivo": (await _last_file(client, conversation)), "hoja": "Ventas",
             "fila_cabecera": 3, "nombre": "otra vista"},
        ),
        text_response("Preparado otra vez."),
    )
    turn = await _send(client, conversation, "prepáralo con otro nombre")
    second = turn["assistant_message"]["artifacts"][0]["dataset_id"]

    canvas = (await client.post("/api/dashboards", json={"name": "Mixto"})).json()["id"]
    for dataset_id in (first, second):
        response = await client.post(
            f"/api/dashboards/{canvas}/visuals",
            json={"dataset_id": dataset_id, "spec": VENTAS_POR_REGION},
        )
        assert response.status_code == 201, response.text

    listed = (await client.get("/api/dashboards")).json()

    assert listed[0]["visual_count"] == 2
    # Un lienzo no pertenece a una fuente: enumera las que usa.
    assert {source["id"] for source in listed[0]["sources"]} == {first, second}


async def _last_file(client: AsyncClient, conversation: str) -> str:
    history = (await client.get(f"/api/conversations/{conversation}/messages")).json()
    for message in history:
        if message["role"] == MessageRole.USER and message["attachments"]:
            return str(message["attachments"][0]["file_id"])
    raise AssertionError("no hay adjuntos en la conversacion")
