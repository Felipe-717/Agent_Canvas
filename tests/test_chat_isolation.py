"""Una conversacion solo conoce lo que ella misma ha preparado.

Enumerarle al modelo todos los conjuntos del usuario tenia dos efectos malos, y
los dos aparecieron en cuanto se probo con datos reales: daba por hecho un
trabajo que no habia hecho ("ya esta preparado", senalando un conjunto que
subio otra pantalla), y mezclaba conversaciones sin relacion.
"""

from __future__ import annotations

from collections.abc import AsyncIterator
from pathlib import Path

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from agentcanvas.application.ports.llm import Role
from agentcanvas.bootstrap.container import Container
from agentcanvas.infrastructure.web.app import create_app
from tests.factories import make_dataset
from tests.fakes import FakeLLM, text_response, tool_response


@pytest.fixture
async def client(container: Container) -> AsyncIterator[AsyncClient]:
    app = create_app(container)
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


@pytest.fixture
def excel(tmp_path: Path) -> bytes:
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Ventas"
    for column, name in enumerate(["region", "valor"], start=1):
        sheet.cell(row=1, column=column, value=name)
    sheet.cell(row=2, column=1, value="Norte")
    sheet.cell(row=2, column=2, value=10)
    path = tmp_path / "ventas.xlsx"
    book.save(path)
    return path.read_bytes()


async def _conversation_with_data(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> tuple[str, str]:
    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("Recibido."))
    turn = (
        await client.post(
            f"/api/conversations/{conversation}/messages",
            data={"text": "toma"},
            files={"file": ("ventas.xlsx", excel, "application/vnd.ms-excel")},
        )
    ).json()
    file_id = turn["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response(
            "preparar_datos", {"archivo": file_id, "hoja": "Ventas", "fila_cabecera": 1}
        ),
        text_response("Preparado."),
    )
    turn = (
        await client.post(
            f"/api/conversations/{conversation}/messages", data={"text": "prepáralo"}
        )
    ).json()
    return conversation, turn["assistant_message"]["artifacts"][0]["dataset_id"]


def _system_notes(llm: FakeLLM) -> str:
    return "\n".join(
        message.content
        for message in llm.requests[-1].messages
        if message.role is Role.SYSTEM
    )


async def test_the_model_is_told_about_data_prepared_in_this_conversation(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation, dataset_id = await _conversation_with_data(client, llm, excel)

    llm.queue(text_response("Sí."))
    await client.post(f"/api/conversations/{conversation}/messages", data={"text": "¿y ahora?"})

    assert dataset_id in _system_notes(llm)


async def test_another_conversation_does_not_see_that_data(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    _, dataset_id = await _conversation_with_data(client, llm, excel)
    otra = (await client.post("/api/conversations")).json()["id"]

    llm.queue(text_response("Hola."))
    await client.post(f"/api/conversations/{otra}/messages", data={"text": "hola"})

    # Sin esto, el modelo cree que ya ha preparado datos que nunca vio.
    assert dataset_id not in _system_notes(llm)


async def test_data_prepared_elsewhere_is_invisible_to_a_conversation(
    client: AsyncClient, llm: FakeLLM, container: Container
) -> None:
    # Lo que exista en el sistema por otra via no es asunto de esta charla.
    ajeno = await make_dataset(
        container,
        name="ajeno",
        csv=b"fecha,region,valor\n2026-01-01,Norte,10\n",
    )

    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("Hola."))
    await client.post(f"/api/conversations/{conversation}/messages", data={"text": "hola"})

    assert ajeno.id not in _system_notes(llm)
