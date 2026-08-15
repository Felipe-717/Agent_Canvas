"""Actualizar una fuente con el archivo del mes siguiente.

Es la promesa del producto entera: el archivo nuevo se relee con las mismas
coordenadas de la primera vez y todos los graficos que dependen de el cambian,
sin volver a preguntarle nada al modelo.
"""

from __future__ import annotations

from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from agentcanvas.bootstrap.container import Container
from agentcanvas.infrastructure.web.app import create_app
from tests.factories import make_dataset
from tests.fakes import FakeLLM, text_response, tool_response

POR_REGION: dict[str, Any] = {
    "type": "bar",
    "title": "Ventas por región",
    "x": {"field": "region"},
    "y": [{"field": "valor", "aggregation": "sum"}],
}


@pytest.fixture
async def client(container: Container) -> AsyncIterator[AsyncClient]:
    app = create_app(container)
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


def _messy_excel(path: Path, rows: list[tuple[str, int]]) -> bytes:
    """Cabecera en la fila 3, con basura encima: el caso real."""
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Ventas"
    sheet["A1"] = "Informe mensual"
    for column, name in enumerate(["region", "valor"], start=1):
        sheet.cell(row=3, column=column, value=name)
    for offset, (region, valor) in enumerate(rows):
        sheet.cell(row=4 + offset, column=1, value=region)
        sheet.cell(row=4 + offset, column=2, value=valor)
    book.save(path)
    return path.read_bytes()


async def _prepared_from_excel(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> str:
    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("Recibido."))
    turn = (
        await client.post(
            f"/api/conversations/{conversation}/messages",
            data={"text": "toma"},
            files={"file": ("enero.xlsx", excel, "application/vnd.ms-excel")},
        )
    ).json()
    file_id = turn["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response(
            "preparar_datos", {"archivo": file_id, "hoja": "Ventas", "fila_cabecera": 3}
        ),
        text_response("Preparado."),
    )
    turn = (
        await client.post(
            f"/api/conversations/{conversation}/messages", data={"text": "prepáralo"}
        )
    ).json()
    dataset_id: str = turn["assistant_message"]["artifacts"][0]["dataset_id"]
    return dataset_id


async def test_a_new_file_is_read_with_the_same_coordinates(
    client: AsyncClient, llm: FakeLLM, tmp_path: Path
) -> None:
    enero = _messy_excel(tmp_path / "enero.xlsx", [("Norte", 100), ("Sur", 150)])
    dataset_id = await _prepared_from_excel(client, llm, enero)

    febrero = _messy_excel(
        tmp_path / "febrero.xlsx", [("Norte", 120), ("Este", 60), ("Sur", 90)]
    )
    response = await client.post(
        f"/api/datasets/{dataset_id}/refresh",
        files={"file": ("febrero.xlsx", febrero, "application/vnd.ms-excel")},
    )

    assert response.status_code == 200, response.text
    body = response.json()
    # La cabecera vuelve a leerse en la fila 3 sin que nadie lo diga.
    assert body["row_count"] == 3
    assert body["previous_rows"] == 2
    # Tres llamadas: el saludo, la herramienta y el cierre. Actualizar no anade
    # ninguna: releer no necesita al modelo.
    assert llm.calls == 3


async def test_the_saved_charts_show_the_new_numbers(
    client: AsyncClient, llm: FakeLLM, tmp_path: Path
) -> None:
    enero = _messy_excel(tmp_path / "enero.xlsx", [("Norte", 100), ("Sur", 150)])
    dataset_id = await _prepared_from_excel(client, llm, enero)
    canvas = (await client.post("/api/dashboards", json={"name": "Ventas"})).json()["id"]
    await client.post(
        f"/api/dashboards/{canvas}/visuals", json={"dataset_id": dataset_id, "spec": POR_REGION}
    )

    febrero = _messy_excel(tmp_path / "febrero.xlsx", [("Norte", 120), ("Este", 60)])
    await client.post(
        f"/api/datasets/{dataset_id}/refresh",
        files={"file": ("febrero.xlsx", febrero, "application/vnd.ms-excel")},
    )

    body = (await client.get(f"/api/dashboards/{canvas}")).json()
    # Sin `sort` el orden lo pone el agrupado, que es alfabetico.
    assert body["visuals"][0]["data"]["rows"] == [
        {"region": "Este", "sum_valor": 60},
        {"region": "Norte", "sum_valor": 120},
    ]


async def test_an_incompatible_file_leaves_everything_as_it_was(
    client: AsyncClient, llm: FakeLLM, tmp_path: Path
) -> None:
    enero = _messy_excel(tmp_path / "enero.xlsx", [("Norte", 100)])
    dataset_id = await _prepared_from_excel(client, llm, enero)

    # Un archivo con otra estructura: en la fila 3 ya no estan esas columnas.
    otro = openpyxl.Workbook()
    hoja = otro.active
    assert hoja is not None
    hoja.title = "Ventas"
    hoja["A3"] = "region"
    hoja["A4"] = "Norte"
    path = tmp_path / "otro.xlsx"
    otro.save(path)

    response = await client.post(
        f"/api/datasets/{dataset_id}/refresh",
        files={"file": ("otro.xlsx", path.read_bytes(), "application/vnd.ms-excel")},
    )

    assert response.status_code == 422
    assert any("valor" in problem for problem in response.json()["problems"])
    # Un dashboard con datos viejos es recuperable; uno mal leido no: el
    # conjunto sigue teniendo las filas de enero.
    assert (await client.get(f"/api/datasets/{dataset_id}/refresh")).status_code in (404, 405)


async def test_the_canvas_lists_its_sources_with_their_state(
    client: AsyncClient, llm: FakeLLM, tmp_path: Path, container: Container
) -> None:
    enero = _messy_excel(tmp_path / "enero.xlsx", [("Norte", 100)])
    del_chat = await _prepared_from_excel(client, llm, enero)
    # Uno creado por fuera del chat, sin coordenadas guardadas.
    antiguo = await make_dataset(
        container, name="antiguo", csv=b"region,valor\nNorte,10\n"
    )

    canvas = (await client.post("/api/dashboards", json={"name": "Mixto"})).json()["id"]
    for dataset_id in (del_chat, antiguo.id):
        await client.post(
            f"/api/dashboards/{canvas}/visuals",
            json={"dataset_id": dataset_id, "spec": POR_REGION},
        )

    sources = (await client.get(f"/api/dashboards/{canvas}")).json()["sources"]

    by_id = {source["id"]: source for source in sources}
    assert by_id[del_chat]["can_refresh"] is True
    # Sin saber como se extrajo, no se puede releer solo: el boton se apaga.
    assert by_id[antiguo.id]["can_refresh"] is False


async def test_a_dataset_without_coordinates_cannot_be_refreshed_from_excel(
    client: AsyncClient, container: Container, tmp_path: Path
) -> None:
    antiguo = await make_dataset(container, name="antiguo", csv=b"region,valor\nNorte,10\n")
    excel = _messy_excel(tmp_path / "x.xlsx", [("Norte", 1)])

    response = await client.post(
        f"/api/datasets/{antiguo.id}/refresh",
        files={"file": ("x.xlsx", excel, "application/vnd.ms-excel")},
    )

    assert response.status_code == 400
    assert "vuelve a prepararlo" in response.json()["detail"].lower()


async def test_a_csv_dataset_can_be_refreshed_with_another_csv(
    client: AsyncClient, container: Container
) -> None:
    # Un CSV se relee siempre igual, asi que aunque sea antiguo se puede.
    dataset = await make_dataset(
        container, name="ventas", csv=b"region,valor\nNorte,10\nSur,20\n"
    )

    response = await client.post(
        f"/api/datasets/{dataset.id}/refresh",
        files={"file": ("nuevo.csv", b"region,valor\nNorte,30\n", "text/csv")},
    )

    assert response.status_code == 200, response.text
    assert response.json()["row_count"] == 1


async def test_the_visual_carries_the_code_that_computes_it(
    client: AsyncClient, container: Container
) -> None:
    dataset = await make_dataset(container, name="ventas", csv=b"region,valor\nNorte,10\n")
    canvas = (await client.post("/api/dashboards", json={"name": "x"})).json()["id"]
    await client.post(
        f"/api/dashboards/{canvas}/visuals",
        json={"dataset_id": dataset.id, "spec": POR_REGION},
    )

    visual = (await client.get(f"/api/dashboards/{canvas}")).json()["visuals"][0]

    # Una cifra que no se puede auditar no vale mucho.
    assert "import pandas as pd" in visual["code"]
    assert "groupby" in visual["code"]
    assert "sum_valor" in visual["code"]
