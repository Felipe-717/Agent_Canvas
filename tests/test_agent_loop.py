"""El bucle de herramientas y el agente que descifra un archivo.

Con el modelo guionado: lo que se comprueba no es que el modelo acierte, sino
que el harness le deje trabajar, le devuelva errores utiles y no le permita
saltarse la validacion.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
import pytest

from agentcanvas.agent.budget import Budget
from agentcanvas.agent.loop import AgentLoop
from agentcanvas.agent.structure_agent import StructureProposal, WorkbookStructureAgent
from agentcanvas.agent.tools import Toolbox, ToolOutcome, tool
from agentcanvas.application.ports.llm import Role
from agentcanvas.infrastructure.tabular.workbook_reader import OpenpyxlWorkbookReader
from tests.fakes import FakeLLM, text_response, tool_response

SCHEMA: dict[str, Any] = {"type": "object", "properties": {"x": {"type": "string"}}}


def _toolbox(outcomes: list[ToolOutcome]) -> tuple[Toolbox, list[dict[str, Any]]]:
    seen: list[dict[str, Any]] = []

    async def handler(arguments: dict[str, Any]) -> ToolOutcome:
        seen.append(arguments)
        return outcomes.pop(0)

    return Toolbox([tool(name="mirar", description="d", parameters=SCHEMA, handler=handler)]), seen


# ------------------------------------------------------------------- toolbox


async def test_an_invented_tool_is_answered_with_the_real_ones() -> None:
    box, _ = _toolbox([])

    outcome = await box.run("borrar_todo", {})

    # Inventarse una herramienta es recuperable: se le dice cuales hay.
    assert "no existe" in outcome.message.lower()
    assert "mirar" in outcome.message
    assert not outcome.done


async def test_a_failing_tool_does_not_break_the_conversation() -> None:
    async def explode(_: dict[str, Any]) -> ToolOutcome:
        raise RuntimeError("la hoja no existe")

    box = Toolbox([tool(name="mirar", description="d", parameters=SCHEMA, handler=explode)])

    outcome = await box.run("mirar", {})

    assert "la hoja no existe" in outcome.message
    assert not outcome.done


# ---------------------------------------------------------------------- bucle


async def test_the_loop_feeds_the_tool_result_back_to_the_model() -> None:
    llm = FakeLLM([tool_response("mirar", {"x": "a"}), text_response("ya lo veo")])
    box, seen = _toolbox([ToolOutcome(message="fila 3: cabeceras")])

    result = await AgentLoop(llm).run([], box)

    assert seen == [{"x": "a"}]
    assert result.text == "ya lo veo"
    # El modelo tiene que haber visto el resultado en el segundo turno.
    tool_messages = [m for m in llm.requests[1].messages if m.role is Role.TOOL]
    assert tool_messages[0].content == "fila 3: cabeceras"


async def test_answering_in_text_is_a_legitimate_ending() -> None:
    # Un agente que explora un archivo caotico debe poder preguntar.
    llm = FakeLLM([text_response("¿Cual de las once hojas te interesa?")])

    result = await AgentLoop(llm).run([], Toolbox([]))

    assert not result.completed
    assert "once hojas" in result.text


async def test_a_terminal_tool_ends_the_loop_with_its_payload() -> None:
    llm = FakeLLM([tool_response("mirar", {})])
    box, _ = _toolbox([ToolOutcome(message="listo", payload={"filas": 85}, done=True)])

    result = await AgentLoop(llm).run([], box)

    assert result.completed
    assert result.payload == {"filas": 85}
    assert llm.exhausted


async def test_running_out_of_budget_still_produces_a_real_message() -> None:
    llm = FakeLLM(
        [
            tool_response("mirar", {}),
            tool_response("mirar", {}),
            text_response("He visto tres hojas con datos. ¿Cual quieres?"),
        ]
    )
    box, _ = _toolbox([ToolOutcome(message="a"), ToolOutcome(message="b")])

    result = await AgentLoop(llm, Budget(max_iterations=2)).run([], box)

    # El trabajo de exploracion ya esta pagado: se aprovecha en vez de soltar
    # un mensaje enlatado.
    assert "tres hojas" in result.text
    assert not result.completed


async def test_the_conversation_comes_back_so_it_can_be_continued() -> None:
    llm = FakeLLM([text_response("¿Que hoja?")])

    result = await AgentLoop(llm).run([], Toolbox([]))

    assert result.messages[-1].role is Role.ASSISTANT
    assert result.messages[-1].content == "¿Que hoja?"


# ---------------------------------------------------------- agente de estructura


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Datos"
    sheet["A1"] = "Informe interno"
    for column, name in enumerate(["Region", "Valor"], start=1):
        sheet.cell(row=3, column=column, value=name)
    sheet.cell(row=4, column=1, value="Norte")
    sheet.cell(row=4, column=2, value=10)
    sheet.cell(row=5, column=1, value="Sur")
    sheet.cell(row=5, column=2, value=20)
    path = tmp_path / "libro.xlsx"
    book.save(path)
    return path


async def test_the_agent_returns_the_extracted_table(workbook: Path, tmp_path: Path) -> None:
    llm = FakeLLM([tool_response("proponer_tabla", {"hoja": "Datos", "fila_cabecera": 3})])
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    result = await agent.inspect(
        workbook, destination=tmp_path / "o.parquet", filename="libro.xlsx"
    )

    assert isinstance(result.payload, StructureProposal)
    assert result.payload.spec.header_row == 3
    assert result.payload.table.row_count == 2
    assert result.payload.table.schema_.column_names == ("region", "valor")


async def test_a_proposal_with_no_data_below_the_header_is_rejected(
    workbook: Path, tmp_path: Path
) -> None:
    llm = FakeLLM(
        [
            # La fila 5 es la ultima con datos: debajo no queda nada que extraer.
            tool_response("proponer_tabla", {"hoja": "Datos", "fila_cabecera": 5}),
            text_response("Me he equivocado de fila"),
        ]
    )
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    result = await agent.inspect(
        workbook, destination=tmp_path / "o.parquet", filename="libro.xlsx"
    )

    assert not result.completed
    # El modelo recibe el motivo, no una excepcion.
    feedback = next(m for m in llm.requests[1].messages if m.role is Role.TOOL)
    assert "ninguna fila" in feedback.content


async def test_pointing_at_a_row_that_does_not_exist_says_so(
    workbook: Path, tmp_path: Path
) -> None:
    llm = FakeLLM(
        [
            tool_response("proponer_tabla", {"hoja": "Datos", "fila_cabecera": 90}),
            text_response("vale"),
        ]
    )
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    await agent.inspect(workbook, destination=tmp_path / "o.parquet", filename="libro.xlsx")

    feedback = next(m for m in llm.requests[1].messages if m.role is Role.TOOL)
    assert "90" in feedback.content
    assert "no existe" in feedback.content


async def test_an_incoherent_proposal_is_explained_not_crashed(
    workbook: Path, tmp_path: Path
) -> None:
    llm = FakeLLM(
        [
            tool_response(
                "proponer_tabla",
                {"hoja": "Datos", "fila_cabecera": 5, "primera_fila_datos": 2},
            ),
            text_response("vale"),
        ]
    )
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    await agent.inspect(workbook, destination=tmp_path / "o.parquet", filename="libro.xlsx")

    feedback = next(m for m in llm.requests[1].messages if m.role is Role.TOOL)
    assert "no es coherente" in feedback.content


async def test_listing_sheets_already_shows_their_first_rows(
    workbook: Path, tmp_path: Path
) -> None:
    llm = FakeLLM([tool_response("listar_hojas", {}), text_response("ok")])
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    await agent.inspect(workbook, destination=tmp_path / "o.parquet", filename="libro.xlsx")

    listing = next(m for m in llm.requests[1].messages if m.role is Role.TOOL).content
    # Adelantar el contenido evita una llamada a `mirar` por cada hoja.
    assert "Datos" in listing
    assert "Region" in listing


async def test_the_user_instruction_reaches_the_model(workbook: Path, tmp_path: Path) -> None:
    llm = FakeLLM([text_response("ok")])
    agent = WorkbookStructureAgent(llm, OpenpyxlWorkbookReader())

    await agent.inspect(
        workbook,
        destination=tmp_path / "o.parquet",
        filename="libro.xlsx",
        instruction="quiero la CAMA 2",
    )

    assert "CAMA 2" in llm.requests[0].messages[-1].content

