"""Lectura de libros que no son tablas.

Las fixtures reproducen patologias tomadas de archivos reales: prosa antes de
la cabecera, fila de ejemplo que hay que descartar, varias tablas lado a lado
en la misma hoja, totales al pie, y hojas auxiliares vacias.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

from agentcanvas.domain.workbook.structure import CSV_SHEET, TableSpec
from agentcanvas.infrastructure.tabular.workbook_reader import OpenpyxlWorkbookReader


@pytest.fixture
def reader() -> OpenpyxlWorkbookReader:
    return OpenpyxlWorkbookReader()


@pytest.fixture
def messy(tmp_path: Path) -> Path:
    """Un libro con las tres patologias mas comunes a la vez."""
    book = openpyxl.Workbook()

    contactos = book.active
    assert contactos is not None
    contactos.title = "Contactos"
    contactos["A1"] = "Seguimiento de contactos 2027"
    contactos["A2"] = "Objetivo: estancia de investigacion"
    contactos["A4"] = "COMO USAR ESTA HOJA"
    contactos["A5"] = "Columnas a llenar TU: prioridad, estado"
    for column, name in enumerate(["Pais", "Institucion", "Prioridad"], start=1):
        contactos.cell(row=7, column=column, value=name)
    contactos.append([])  # deja el cursor lejos; se escribe por coordenadas
    for offset, row in enumerate(
        [
            ["Ejemplo", "(borra esta fila)", "A"],
            ["Italia", "Sapienza", "A"],
            ["Francia", "Inria", "B"],
            ["Italia", "Politecnico", "A"],
        ]
    ):
        for column, value in enumerate(row, start=1):
            contactos.cell(row=8 + offset, column=column, value=value)

    camas = book.create_sheet("Camas")
    camas["A1"] = "CAMAS DE GERMINACION"
    for start, nombre in [(1, "CAMA 1"), (4, "CAMA 2")]:
        camas.cell(row=2, column=start, value=nombre)
        for offset, name in enumerate(["Fecha", "Semilla", "Cantidad"]):
            camas.cell(row=3, column=start + offset, value=name)
    camas.cell(row=4, column=1, value="2026-01-10")
    camas.cell(row=4, column=2, value="otobo")
    camas.cell(row=4, column=3, value=144)
    camas.cell(row=5, column=1, value="2026-01-11")
    camas.cell(row=5, column=2, value="azuceno")
    camas.cell(row=5, column=3, value=60)
    camas.cell(row=4, column=4, value="2026-02-01")
    camas.cell(row=4, column=5, value="balso")
    camas.cell(row=4, column=6, value=930)
    camas.cell(row=6, column=1, value="TOTAL")
    camas.cell(row=6, column=3, value=204)

    book.create_sheet("Hoja3")  # vacia, como las que deja Excel

    path = tmp_path / "raro.xlsx"
    book.save(path)
    return path


# ------------------------------------------------------------------ exploracion


def test_the_overview_lists_every_sheet_with_its_size(
    reader: OpenpyxlWorkbookReader, messy: Path
) -> None:
    overview = reader.overview(messy)

    assert overview.sheet_names == ("Contactos", "Camas", "Hoja3")
    contactos = overview.get("Contactos")
    assert contactos is not None
    assert contactos.filled_cells > 0


def test_empty_sheets_drop_out_of_the_candidates(
    reader: OpenpyxlWorkbookReader, messy: Path
) -> None:
    # Sin esto, el agente gastaria iteraciones mirando 'Hoja3'.
    candidates = reader.overview(messy).candidates

    assert "Hoja3" not in [sheet.name for sheet in candidates]
    # Y las que quedan van de mas a menos prometedora.
    assert candidates[0].filled_cells >= candidates[-1].filled_cells


def test_peeking_shows_the_real_row_numbers(
    reader: OpenpyxlWorkbookReader, messy: Path
) -> None:
    window = reader.peek(messy, sheet="Contactos", first_row=6, rows=3, columns=3)

    rendered = window.render()
    # Los numeros de fila son justo lo que el agente necesita para poder decir
    # "la cabecera esta en la fila 7".
    assert "   7 | Pais" in rendered
    assert window.first_row == 6


def test_empty_cells_are_visible_as_placeholders(
    reader: OpenpyxlWorkbookReader, messy: Path
) -> None:
    rendered = reader.peek(messy, sheet="Contactos", first_row=1, rows=2, columns=3).render()
    # Un hueco tiene que verse; si no, el modelo no distingue "vacio" de
    # "columna que no existe".
    assert "·" in rendered


def test_peeking_an_unknown_sheet_says_which_ones_hay(
    reader: OpenpyxlWorkbookReader, messy: Path
) -> None:
    with pytest.raises(ValueError, match="Contactos"):
        reader.peek(messy, sheet="Inexistente")


# ------------------------------------------------------------------- extraccion


def test_the_header_can_live_far_from_the_first_row(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    spec = TableSpec(sheet="Contactos", header_row=7)

    table = reader.extract(messy, spec, destination=tmp_path / "out.parquet")

    assert table.schema_.column_names == ("pais", "institucion", "prioridad")
    assert table.row_count == 4  # incluye todavia la fila de ejemplo


def test_the_example_row_can_be_skipped(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    spec = TableSpec(sheet="Contactos", header_row=7, skip_rows=(8,))

    table = reader.extract(messy, spec, destination=tmp_path / "out.parquet")

    assert table.row_count == 3
    assert table.preview[0]["pais"] == "Italia"


def test_two_tables_side_by_side_are_read_separately(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    # La patologia que ningun lector ingenuo sobrevive.
    primera = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, last_data_row=5, first_column=1, last_column=3),
        destination=tmp_path / "a.parquet",
    )
    segunda = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, last_data_row=5, first_column=4, last_column=6),
        destination=tmp_path / "b.parquet",
    )

    assert primera.row_count == 2
    assert segunda.row_count == 1
    assert segunda.preview[0]["semilla"] == "balso"


def test_the_totals_row_can_be_left_out(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    con_total = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, first_column=1, last_column=3),
        destination=tmp_path / "a.parquet",
    )
    sin_total = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, last_data_row=5, first_column=1, last_column=3),
        destination=tmp_path / "b.parquet",
    )

    # Sin el corte, "TOTAL" entra como si fuese una siembra mas y duplica la suma.
    assert con_total.row_count == 3
    assert sin_total.row_count == 2


def test_columns_without_a_header_still_get_a_name(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet["A1"] = "nombre"
    sheet["C1"] = "total"  # B1 vacia, como al deshacer celdas combinadas
    sheet["A2"] = "x"
    sheet["B2"] = "dato suelto"
    sheet["C2"] = 5
    path = tmp_path / "hueco.xlsx"
    book.save(path)

    table = reader.extract(
        path, TableSpec(sheet=sheet.title, header_row=1), destination=tmp_path / "o.parquet"
    )

    assert table.schema_.column_names == ("nombre", "columna_2", "total")


def test_a_csv_looks_like_a_workbook_with_one_sheet(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    # Asi el resto del sistema no tiene que distinguir CSV de Excel.
    source = tmp_path / "v.csv"
    source.write_text("basura\n\nfecha;region;valor\n2026-01-01;Norte;10\n", encoding="utf-8")

    overview = reader.overview(source)
    table = reader.extract(
        source, TableSpec(sheet=CSV_SHEET, header_row=3), destination=tmp_path / "o.parquet"
    )

    assert overview.sheet_names == (CSV_SHEET,)
    assert table.schema_.column_names == ("fecha", "region", "valor")
    assert table.row_count == 1


# --------------------------------------------------------------------- dominio


def test_a_spec_whose_data_starts_before_the_header_is_rejected() -> None:
    with pytest.raises(ValueError, match="despues de la fila de cabecera"):
        TableSpec(sheet="x", header_row=5, first_data_row=3)


def test_an_inverted_column_range_is_rejected() -> None:
    with pytest.raises(ValueError, match="columnas esta invertido"):
        TableSpec(sheet="x", header_row=1, first_column=6, last_column=2)


def test_the_description_is_readable_for_a_human() -> None:
    spec = TableSpec(sheet="Camas", header_row=3, last_data_row=11, first_column=4, last_column=6)

    described = spec.describe()

    assert "Camas" in described
    assert "cabecera en la fila 3" in described
    assert "columnas 4-6" in described


# --------------------------------------------------------------------- avisos


def test_side_by_side_tables_are_reported_not_silently_merged(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    # Extraer las dos camas de golpe no falla: produce una tabla de seis
    # columnas que parece correcta. Callarselo es lo que genera graficos que
    # solo cubren la primera tabla sin que nadie se entere.
    table = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, last_data_row=5, first_column=1, last_column=6),
        destination=tmp_path / "o.parquet",
    )

    assert table.schema_.column_names == (
        "fecha", "semilla", "cantidad", "fecha_2", "semilla_2", "cantidad_2",
    )
    aviso = next(w for w in table.warnings if "repite" in w)
    assert "2 veces" in aviso
    assert "primera_columna" in aviso


def test_a_single_table_raises_no_alarm(
    reader: OpenpyxlWorkbookReader, messy: Path, tmp_path: Path
) -> None:
    table = reader.extract(
        messy,
        TableSpec(sheet="Camas", header_row=3, last_data_row=5, first_column=1, last_column=3),
        destination=tmp_path / "o.parquet",
    )

    assert not any("repite" in warning for warning in table.warnings)


def test_rows_with_an_empty_first_column_are_reported(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    # Una fila de totales entra como si fuera un dato mas y encabeza cualquier
    # "top 10". Paso de verdad con un inventario real.
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    for column, name in enumerate(["especie", "cantidad"], start=1):
        sheet.cell(row=1, column=column, value=name)
    sheet.cell(row=2, column=1, value="Aguacatillo")
    sheet.cell(row=2, column=2, value=422)
    sheet.cell(row=3, column=2, value=3452)  # total, sin nombre
    path = tmp_path / "inventario.xlsx"
    book.save(path)

    table = reader.extract(
        path, TableSpec(sheet=sheet.title, header_row=1), destination=tmp_path / "o.parquet"
    )

    aviso = next(w for w in table.warnings if "vacio" in w)
    assert "1 fila" in aviso
    assert "ultima_fila_datos" in aviso


def test_a_totals_row_no_longer_breaks_the_extraction(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    # La palabra TOTAL en una columna de fechas hacia que Parquet se negara a
    # escribir el archivo entero, con un error de Arrow que no le decia nada a
    # nadie. Leerla como texto y avisar es recuperable.
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    for column, name in enumerate(["fecha", "cantidad"], start=1):
        sheet.cell(row=1, column=column, value=name)
    # Una fecha de verdad: openpyxl devuelve datetime, y ahi es donde choca
    # con el texto de la fila de totales.
    sheet.cell(row=2, column=1, value=datetime(2026, 1, 10))
    sheet.cell(row=2, column=2, value=144)
    sheet.cell(row=3, column=1, value="TOTAL")
    sheet.cell(row=3, column=2, value=144)
    path = tmp_path / "con_total.xlsx"
    book.save(path)

    table = reader.extract(
        path, TableSpec(sheet=sheet.title, header_row=1), destination=tmp_path / "o.parquet"
    )

    assert table.row_count == 2
    assert any("mezclan varios tipos" in warning for warning in table.warnings)


def test_a_repeated_group_is_detected_despite_small_variations(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    # En la hoja real, una cama ponia "chapola" donde las demas ponen "semilla".
    # Exigir un patron perfecto hacia que el aviso no saltara justo ahi.
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    cabecera = ["fecha", "semilla", "cantidad", "fecha", "chapola", "cantidad"]
    for column, name in enumerate(cabecera, start=1):
        sheet.cell(row=1, column=column, value=name)
    for column in range(1, 7):
        sheet.cell(row=2, column=column, value=column)
    path = tmp_path / "variado.xlsx"
    book.save(path)

    table = reader.extract(
        path, TableSpec(sheet=sheet.title, header_row=1), destination=tmp_path / "o.parquet"
    )

    aviso = next(w for w in table.warnings if "repite" in w)
    assert "fecha" in aviso and "cantidad" in aviso
    # Y explica donde vive la identidad de cada bloque, que es lo que el modelo
    # no adivinaba: buscaba "CAMA 2" como si fuera un valor de los datos.
    assert "fila de encima" in aviso


def test_an_iso_date_keeps_its_day_and_month(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    """El 5 de marzo no puede convertirse en el 3 de mayo.

    Las fechas de fuera se leen con `dayfirst`, que es como se escriben en
    espanol, pero aplicarlo tambien a las ISO intercambiaba dia y mes sin decir
    nada: la columna seguia siendo de fechas, los graficos seguian saliendo y
    todo cuadraba salvo que estaba mal. Este es el peor tipo de fallo que puede
    tener el sistema, asi que se fija con un test.
    """
    source = tmp_path / "iso.csv"
    source.write_text("fecha,valor\n2026-03-05,10\n2026-07-11,20\n", encoding="utf-8")

    table = reader.extract(
        source, TableSpec(sheet=CSV_SHEET, header_row=1), destination=tmp_path / "o.parquet"
    )

    assert table.schema_.get("fecha") is not None
    assert [str(row["fecha"])[:10] for row in table.preview] == ["2026-03-05", "2026-07-11"]


def test_a_day_first_date_is_still_read_day_first(
    reader: OpenpyxlWorkbookReader, tmp_path: Path
) -> None:
    # Y lo de arriba no puede haberse llevado por delante el caso corriente:
    # "05/03/2026" escrito por alguien de aqui es el 5 de marzo.
    source = tmp_path / "local.csv"
    source.write_text("fecha,valor\n05/03/2026,10\n11/07/2026,20\n", encoding="utf-8")

    table = reader.extract(
        source, TableSpec(sheet=CSV_SHEET, header_row=1), destination=tmp_path / "o.parquet"
    )

    assert [str(row["fecha"])[:10] for row in table.preview] == ["2026-03-05", "2026-07-11"]
