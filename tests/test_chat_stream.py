"""El chat contando lo que hace mientras lo hace.

Explorar un libro de once hojas lleva sus segundos. Sin senal, la espera se lee
como que algo se ha colgado; con ella, como trabajo en curso.
"""

from __future__ import annotations

import json
from collections.abc import AsyncIterator
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from httpx import ASGITransport, AsyncClient

from agentcanvas.agent.tools import Toolbox, ToolOutcome, tool
from agentcanvas.bootstrap.container import Container
from agentcanvas.infrastructure.web.app import create_app
from tests.fakes import FakeLLM, text_response, tool_response

VENTAS_POR_REGION: dict[str, Any] = {
    "type": "bar",
    "title": "Ventas por región",
    "x": {"field": "region"},
    "y": [{"field": "valor", "aggregation": "sum"}],
}


@pytest.fixture
async def client(container: Container) -> AsyncIterator[AsyncClient]:
    app = create_app(container)
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as c:
        yield c


@pytest.fixture
def excel(tmp_path: Path) -> bytes:
    book = openpyxl.Workbook()
    sheet = book.active
    assert sheet is not None
    sheet.title = "Ventas"
    for column, name in enumerate(["region", "valor"], start=1):
        sheet.cell(row=1, column=column, value=name)
    sheet.cell(row=2, column=1, value="Norte")
    sheet.cell(row=2, column=2, value=10)
    path = tmp_path / "ventas.xlsx"
    book.save(path)
    return path.read_bytes()


async def _stream(
    client: AsyncClient, conversation: str, text: str, upload: bytes | None = None
) -> list[tuple[str, dict[str, Any]]]:
    files = {"file": ("ventas.xlsx", upload, "application/vnd.ms-excel")} if upload else None
    async with client.stream(
        "POST",
        f"/api/conversations/{conversation}/messages/stream",
        data={"text": text},
        files=files,
    ) as response:
        assert response.status_code == 200, await response.aread()
        assert response.headers["content-type"].startswith("text/event-stream")
        body = "".join([chunk async for chunk in response.aiter_text()])
    return _parse(body)


def _parse(body: str) -> list[tuple[str, dict[str, Any]]]:
    events: list[tuple[str, dict[str, Any]]] = []
    for block in body.strip().split("\n\n"):
        lines = dict(line.split(": ", 1) for line in block.splitlines() if ": " in line)
        if "event" in lines and "data" in lines:
            events.append((lines["event"], json.loads(lines["data"])))
    return events


# ------------------------------------------------------------------ etiquetas


def test_a_tool_describes_itself_with_its_arguments() -> None:
    async def handler(_: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(message="")

    box = Toolbox(
        [
            tool(
                name="mirar",
                description="d",
                parameters={"type": "object"},
                handler=handler,
                activity="Mirando la hoja {hoja}",
            )
        ]
    )

    assert box.describe("mirar", {"hoja": "INVENTARIO"}) == "Mirando la hoja INVENTARIO"


def test_a_missing_argument_does_not_break_the_message() -> None:
    async def handler(_: dict[str, Any]) -> ToolOutcome:
        return ToolOutcome(message="")

    box = Toolbox(
        [
            tool(
                name="mirar",
                description="d",
                parameters={"type": "object"},
                handler=handler,
                activity="Mirando la hoja {hoja}",
            )
        ]
    )

    # Un aviso de progreso jamas puede tumbar un turno.
    assert box.describe("mirar", {}) == "Mirando la hoja …"


# -------------------------------------------------------------------- stream


async def test_the_stream_narrates_each_tool_and_ends_with_the_turn(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("Recibido."))
    turn = await _stream(client, conversation, "toma", upload=excel)
    file_id = turn[-1][1]["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response("listar_hojas", {"archivo": file_id}, call_id="c1"),
        tool_response(
            "preparar_datos",
            {"archivo": file_id, "hoja": "Ventas", "fila_cabecera": 1},
            call_id="c2",
        ),
        text_response("Preparado."),
    )

    events = await _stream(client, conversation, "prepáralo")

    activities = [payload["text"] for name, payload in events if name == "activity"]
    assert activities == ["Abriendo el archivo", "Preparando los datos de Ventas"]
    assert events[-1][0] == "turn"
    assert events[-1][1]["assistant_message"]["artifacts"][0]["kind"] == "dataset"


async def test_a_plain_answer_streams_no_activity(
    client: AsyncClient, llm: FakeLLM
) -> None:
    # Conversar no es trabajar: no hay nada que narrar.
    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("Puedo ayudarte con hojas de cálculo."))

    events = await _stream(client, conversation, "hola")

    assert [name for name, _ in events] == ["turn"]
    assert events[0][1]["assistant_message"]["text"].startswith("Puedo ayudarte")


async def test_the_activity_arrives_before_the_tool_finishes(
    client: AsyncClient, llm: FakeLLM, excel: bytes
) -> None:
    conversation = (await client.post("/api/conversations")).json()["id"]
    llm.queue(text_response("ok"))
    turn = await _stream(client, conversation, "toma", upload=excel)
    file_id = turn[-1][1]["user_message"]["attachments"][0]["file_id"]

    llm.queue(
        tool_response("mirar", {"archivo": file_id, "hoja": "Ventas"}),
        text_response("Ya lo veo."),
    )
    events = await _stream(client, conversation, "mira la hoja Ventas")

    # El aviso nombra la hoja, que es lo unico que le interesa a quien espera.
    assert ("activity", {"text": "Mirando la hoja Ventas"}) in events


async def test_a_late_failure_travels_as_an_event(
    client: AsyncClient, llm: FakeLLM
) -> None:
    conversation = (await client.post("/api/conversations")).json()["id"]

    events = await _stream(client, "conv_inexistente", "hola")

    # La cabecera ya viajo: un 404 no cabe, asi que va como evento.
    assert events[-1][0] == "error"
    assert events[-1][1]["error"] == "NotFoundError"
    assert conversation not in json.dumps(events)
