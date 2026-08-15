"""Lectura exploratoria de libros de Excel y CSV.

Trabaja con celdas, no con tablas. Es lo que permite que el agente mire un
archivo antes de suponer nada sobre el, que es exactamente lo que hace falta
cuando la cabecera esta en la fila 11 y hay tres tablas en la misma hoja.
"""

from __future__ import annotations

import csv
import re
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd
from pandas.api import types as pdt

from agentcanvas.application.ports.tabular import NormalizedTable
from agentcanvas.domain.workbook.structure import (
    CSV_SHEET,
    CellWindow,
    SheetOverview,
    TableSpec,
    WorkbookOverview,
)
from agentcanvas.infrastructure.tabular.normalize import (
    finalize,
    logical_type,
    preview_rows_of,
)

_CSV_ENCODINGS = ("utf-8-sig", "utf-8", "cp1252", "latin-1")
_MAX_NAMED_COLUMNS = 4
"""Cuantas columnas se nombran en un aviso antes de resumir."""

_MAX_SCAN_ROWS = 2000
"""Tope al medir densidad: un libro de 100k filas no necesita recorrerse entero
para saber si su primera hoja tiene pinta de tabla."""


class OpenpyxlWorkbookReader:
    """Implementa `WorkbookReaderPort`."""

    # ------------------------------------------------------------- exploracion

    def overview(self, source: Path) -> WorkbookOverview:
        if _is_csv(source):
            rows = _read_csv_cells(source)
            filled = sum(1 for row in rows for cell in row if cell)
            width = max((len(row) for row in rows), default=0)
            return WorkbookOverview(
                sheets=(
                    SheetOverview(
                        name=CSV_SHEET,
                        rows=len(rows),
                        columns=width,
                        filled_cells=filled,
                        is_empty=not filled,
                    ),
                )
            )

        book = openpyxl.load_workbook(source, read_only=True, data_only=True)
        try:
            sheets = tuple(_overview_of(book[name]) for name in book.sheetnames)
        finally:
            book.close()
        return WorkbookOverview(sheets=sheets)

    def peek(
        self,
        source: Path,
        *,
        sheet: str,
        first_row: int = 1,
        rows: int = 15,
        first_column: int = 1,
        columns: int = 12,
    ) -> CellWindow:
        last_row = first_row + rows - 1
        last_column = first_column + columns - 1
        grid = _cells(source, sheet, first_row, last_row, first_column, last_column)
        return CellWindow(
            sheet=sheet,
            first_row=first_row,
            first_column=first_column,
            rows=tuple(tuple(_text(cell) for cell in row) for row in grid),
        )

    # -------------------------------------------------------------- extraccion

    def extract(self, source: Path, spec: TableSpec, *, destination: Path) -> NormalizedTable:
        header = _cells(
            source,
            spec.sheet,
            spec.header_row,
            spec.header_row,
            spec.first_column,
            spec.last_column,
        )
        if not header:
            raise ValueError(f"La fila {spec.header_row} no existe en la hoja '{spec.sheet}'")

        body = _cells(
            source,
            spec.sheet,
            spec.data_start,
            spec.last_data_row,
            spec.first_column,
            spec.last_column,
        )
        skipped = {row - spec.data_start for row in spec.skip_rows}
        body = [row for index, row in enumerate(body) if index not in skipped]

        names = _header_names(header[0])
        width = len(names)
        frame = pd.DataFrame(
            [list(row[:width]) + [None] * max(0, width - len(row)) for row in body],
            columns=pd.Index(names),
        )
        frame = _clean(frame, spec)
        if frame.empty or frame.columns.empty:
            # El mensaje viaja hasta el modelo en el ciclo de correccion, asi
            # que dice que hacer, no que ha fallado por dentro.
            raise ValueError(
                f"Con esas coordenadas no sale ninguna fila de datos en la hoja "
                f"'{spec.sheet}'. Revisa la fila de cabecera y el rango de filas."
            )
        return _normalize(frame, destination)


# ---------------------------------------------------------------- utilidades


def _is_csv(source: Path) -> bool:
    return source.suffix.lower() in (".csv", ".txt", ".tsv")


def _overview_of(sheet: Any) -> SheetOverview:
    filled = 0
    scanned = 0
    for row in sheet.iter_rows(max_row=_MAX_SCAN_ROWS, values_only=True):
        scanned += 1
        filled += sum(1 for cell in row if cell is not None and str(cell).strip())
    rows = sheet.max_row or 0
    columns = sheet.max_column or 0
    # Si se dejo de contar antes del final, se extrapola: solo se usa para
    # ordenar hojas por interes, no para decidir nada irreversible.
    if rows > scanned and scanned:
        filled = int(filled * rows / scanned)
    return SheetOverview(
        name=sheet.title,
        rows=rows,
        columns=columns,
        filled_cells=filled,
        is_empty=filled == 0,
    )


def _cells(
    source: Path,
    sheet: str,
    first_row: int,
    last_row: int | None,
    first_column: int,
    last_column: int | None,
) -> list[list[Any]]:
    if _is_csv(source):
        rows = _read_csv_cells(source)
        end = last_row or len(rows)
        selected = rows[first_row - 1 : end]
        return [
            list(row[first_column - 1 : last_column if last_column else None])
            for row in selected
        ]

    book = openpyxl.load_workbook(source, read_only=True, data_only=True)
    try:
        if sheet not in book.sheetnames:
            raise ValueError(f"La hoja '{sheet}' no existe. Hay: {', '.join(book.sheetnames)}")
        worksheet = book[sheet]
        return [
            list(row)
            for row in worksheet.iter_rows(
                min_row=first_row,
                max_row=last_row,
                min_col=first_column,
                max_col=last_column,
                values_only=True,
            )
        ]
    finally:
        book.close()


def _read_csv_cells(source: Path) -> list[list[str]]:
    for encoding in _CSV_ENCODINGS:
        try:
            text = source.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
        lines = text.splitlines()
        return [list(row) for row in csv.reader(lines, delimiter=_delimiter_of(lines))]
    raise ValueError(f"No se pudo decodificar '{source.name}'")


def _delimiter_of(lines: list[str]) -> str:
    """Elige el separador por el que mas columnas produce de forma consistente.

    `csv.Sniffer` se apoya en las primeras lineas, y en un archivo exportado a
    mano las primeras lineas suelen ser un titulo suelto sin separador alguno,
    con lo que adivina mal. Contar columnas sobre todo el archivo es tosco pero
    no se deja enganar por la basura de cabecera.
    """
    best, best_score = ",", 0.0
    for candidate in (",", ";", "\t", "|"):
        counts = [len(row) for row in csv.reader(lines[:200], delimiter=candidate) if row]
        if not counts:
            continue
        # Se premia el ancho tipico, no el maximo: una linea larga suelta no
        # debe decidir por todo el archivo.
        typical = sorted(counts)[len(counts) // 2]
        if typical > best_score:
            best, best_score = candidate, typical
    return best


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()[:19]
    return str(value).strip()


def _header_names(row: list[Any]) -> list[str]:
    """Nombres de columna, rellenando los huecos.

    Una cabecera con celdas vacias es normal cuando venia de celdas combinadas;
    dejarlas sin nombre haria imposible referirse a esas columnas.
    """
    names: list[str] = []
    for index, value in enumerate(row, start=1):
        text = _text(value).replace("\n", " ").strip()
        names.append(text or f"columna_{index}")
    return names


def _clean(frame: pd.DataFrame, spec: TableSpec) -> pd.DataFrame:
    if spec.drop_empty_rows:
        frame = frame.dropna(how="all")
    if spec.drop_empty_columns:
        frame = frame.dropna(axis=1, how="all")
    return frame.reset_index(drop=True)


def _normalize(frame: pd.DataFrame, destination: Path) -> NormalizedTable:
    """Deja la tabla en forma canonica y la escribe en Parquet.

    Reutiliza la normalizacion del lector simple para que un archivo limpio y
    uno rescatado de un Excel caotico acaben siendo indistinguibles.
    """
    from agentcanvas.domain.dataset.schema import ColumnSchema, DatasetSchema

    frame = finalize(frame)
    frame, mixed = _unify_mixed_columns(frame)
    destination.parent.mkdir(parents=True, exist_ok=True)
    frame.to_parquet(destination, index=False)

    columns = tuple(
        ColumnSchema(
            name=str(name),
            original_name=str(original),
            type=logical_type(frame[name]),
            nullable=bool(frame[name].isna().any()),
        )
        for name, original in zip(frame.columns, frame.attrs["original_names"], strict=True)
    )
    return NormalizedTable(
        schema_=DatasetSchema(columns=columns),
        row_count=len(frame),
        preview=preview_rows_of(frame, 10),
        warnings=_warnings(frame) + mixed,
    )


def _unify_mixed_columns(frame: pd.DataFrame) -> tuple[pd.DataFrame, tuple[str, ...]]:
    """Pasa a texto las columnas que mezclan tipos.

    Una fila de totales pone la palabra TOTAL en una columna de fechas, y
    entonces Parquet se niega a escribir el archivo entero. Antes eso rompia la
    extraccion con un error de Arrow que no le dice nada a nadie; ahora la
    columna se lee como texto y se avisa, que es recuperable.
    """
    afectadas: list[str] = []
    for column in frame.columns:
        series = frame[column]
        if not pdt.is_object_dtype(series):
            continue
        present = series.dropna()
        if present.empty or len({type(value) for value in present}) == 1:
            continue
        frame[column] = series.map(lambda value: None if pd.isna(value) else str(value))
        afectadas.append(str(column))

    if not afectadas:
        return frame, ()
    # Un aviso por columna serian veinte lineas en una hoja ancha: mucho
    # contexto y ninguna informacion nueva a partir de la segunda.
    nombres = ", ".join(afectadas[:_MAX_NAMED_COLUMNS])
    resto = len(afectadas) - _MAX_NAMED_COLUMNS
    if resto > 0:
        nombres += f" y {resto} mas"
    return frame, (
        f"{len(afectadas)} columna(s) mezclan varios tipos de dato y se han leido "
        f"como texto ({nombres}). Suele pasar cuando una fila de totales o un "
        f"encabezado se cuela entre los datos; si es el caso, acota "
        f"`ultima_fila_datos`.",
    )


def _warnings(frame: pd.DataFrame) -> tuple[str, ...]:
    """Lo que huele raro en una extraccion que no ha fallado."""
    avisos: list[str] = []

    repetido = _repeated_group([str(column) for column in frame.columns])
    if repetido is not None:
        grupo, veces = repetido
        avisos.append(
            f"La cabecera repite ({', '.join(grupo)}) hasta {veces} veces. Casi "
            f"seguro son {veces} tablas puestas una al lado de otra, no una sola: "
            f"cada bloque de columnas es una tabla distinta, y su identidad esta "
            f"en la fila de encima, no dentro de los datos. Prepara UNA sola "
            f"acotando `primera_columna` y `ultima_columna`, y pregunta al usuario "
            f"cual quiere si no lo ha dicho."
        )

    if len(frame.columns) > 0:
        primera = frame.columns[0]
        huecos = int(frame[primera].isna().sum())
        if huecos:
            avisos.append(
                f"Hay {huecos} fila(s) con '{primera}' vacio. Suelen ser totales o "
                f"separadores, y falsean cualquier suma o maximo. Si lo son, "
                f"vuelve a preparar acotando `ultima_fila_datos`."
            )
    return tuple(avisos)


def _repeated_group(names: list[str]) -> tuple[tuple[str, ...], int] | None:
    """Nombres de columna que se repiten, si delatan tablas puestas en paralelo.

    No se busca periodicidad exacta. Las cabeceras reales tienen variaciones -en
    una hoja con nueve camas de germinacion, una ponia `chapola` donde las demas
    ponen `semilla`- y exigir un patron perfecto hacia que no saltara el aviso
    justo en el caso que mas importaba.
    """
    bases = [re.sub(r"_\d+$", "", name) for name in names]
    counts = Counter(bases)
    repeated = {name: n for name, n in counts.items() if n > 1}
    if not repeated:
        return None

    veces = max(repeated.values())
    # Dos columnas con el mismo nombre son una cabecera duplicada. Que se repita
    # un grupo entero, o que una sola se repita tres veces, ya es otra cosa.
    if veces < 3 and len(repeated) < 2:
        return None
    grupo = tuple(dict.fromkeys(name for name in bases if name in repeated))
    return grupo, veces
